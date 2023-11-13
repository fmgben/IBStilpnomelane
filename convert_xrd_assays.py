import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import Workbook
import re
import numpy as np
wb = Workbook()
from openpyxl import load_workbook
## import the new bv assay data
wb = load_workbook(filename = r'data\Assay\4471 DTR Assay data.R2.xlsx')
sheets = wb.sheetnames
wb.close()

loi_map = {'LOI371-650':'LOI650',
            'LOI650-1000':'LOI1000',
            'LOI1000': 'LOI',
            'LOITotal':'LOI',
            'LOI110-425':'LOI371','LOI425-650':'LOI650'}


column_map = {'Unnamed: 1':'BHID','Unnamed: 2':'Composite', 'Unnamed: 3':'Size', 'Unnamed: 4':'Type','Pruduct':'Product'}
tmp_concat = []

for i in sheets:
    tmp_columns = pd.read_excel(r'data\Assay\4471 DTR Assay data.R2.xlsx',sheet_name=i,skiprows=1,nrows=1)
    tmp_data =pd.read_excel(r'data\Assay\4471 DTR Assay data.R2.xlsx',sheet_name=i,skiprows=2)
    ncolumns = len(tmp_columns.columns)
    new_cols = {}
    for zc in zip(tmp_data.columns[0:ncolumns],tmp_columns.columns[0:ncolumns].to_list()):
        new_cols.update({zc[0]:zc[1]})
    tmp_data.rename(columns=new_cols,inplace=True)
    tmp_data.rename(columns=column_map,inplace=True)
    tmp_data.rename(columns=loi_map,inplace=True)
    numsample = tmp_data.isna().sum(1) <6
    tmp_concat.append(tmp_data[numsample])


data = pd.concat(tmp_concat,axis=0)
good_columns = [c for c in data.columns if not c.startswith('Unnamed')]
data = data[good_columns].copy()

# clean up all the data

reg = re.compile('TAILS|CONS|FEED')
def reg_cleaner(x):
    if isinstance(x, str): 
        yr = reg.findall(x)
        if len(yr)>0:
            y = yr[0]
        else:
            y = ''
    else:
        y = ''
    return y

data.Type = data.Type.str.upper().map(reg_cleaner)


# remove some issues with excel formatted data
data.Type =data.Type.str.replace("R-'DTR",'R-DTR')
# clean out leading and trailing white space
data.Size = data.Size.str.strip()
data.Composite = data.Composite.str.strip()
# Replace the XRF
data.columns = [c.replace(' XRF','') for c in data.columns]
data.to_csv(r'reference\compiled_assays.csv', index=False)

batch2 = pd.read_csv('reference\Batch2_XRD_Assay_From_BBWi.csv')

batch2.columns = [c.strip() for c in batch2.columns]
def clean_lt(x):
    if isinstance(x, str):
        y =x.strip().replace('<','-')
    else:
        y = x
    return y
batch2 = batch2.applymap(clean_lt)

batch2 = batch2.rename(columns=loi_map)
batch2['BHID'] = batch2.Composite.map(lambda x: x.split('-')[0])
batch2['CSS (µm)'] =''
batch2['Type'] = 'FEED'
batch2.Composite = batch2.Composite.map(lambda x:f'Comp {x.split("-")[2]}')
data = pd.concat([data, batch2])

### ===== xrd===== 

def clean_size(x):
    if isinstance(x,str):
        y = x.strip().replace('-','').replace('um','')
    else:
        y = ''
    return y

def clean_type(x):
    if isinstance(x,str):
        y = reg.findall(x)[0]
    else:
        y = ''
    return y

def clean_mineral(x):
    x = x.split('-')
    y = '{} Comp {}'.format(x[0],x[2]).strip()
    return y

def float_to_str_size(x):
    if isinstance(x, str):
        y = ''
    elif isinstance(x, float):
        y = str(int(x))
    elif isinstance(x,np.nan):
        y = str(int(x))

    return y
column_map_xrd = {'Unnamed: 0':'BHID','Unnamed: 1':'Composite', 'Unnamed: 3':'Size', 'Unnamed: 4':'Type','Pruduct':'Product'}

qxrd = pd.read_excel(r'data\XRD\XRD\QXRD_Summary_20211129 FS.xlsx',sheet_name=2)
qxrd.rename(columns=column_map_xrd, inplace=True)
qxrd['Product'] = qxrd['Product'].str.upper()

qxrd1 = pd.read_csv(r'reference\AN0767- QXRD UNROUNDED data.csv')
qxrd1.rename(columns={'Spinel group - Magnetite':'Spinel group','Calcite group':'Calcite group - Siderite','Hematite':'Hematite group','Calcite group - Calcite':'Calcite group'},inplace=True)
# rename some of the columns that we are missing or are mispeled


qxrd1.Mineral = qxrd1.Mineral.map(clean_mineral)
qxrd1['Sample'] =qxrd1.Mineral
qxrd1['Composite'] = qxrd1.Sample.map(lambda x:' '.join(x.split(' ')[1:]))
qxrd1['BHID'] = qxrd1.Mineral.map(lambda x: x.split(' ')[0])
qxrd1['Product'] = 'FEED'
qxrd = pd.concat([qxrd,qxrd1],axis=0)

qxrd.loc[qxrd['CSS (µm)'].isna(),'CSS (µm)'] =''
qxrd['CSS (µm)'] = qxrd['CSS (µm)'].map(lambda x: float_to_str_size(x))

qxrd['CompositeID'] = qxrd.BHID.str.strip()+' '+qxrd.Composite.str.strip()+' '+qxrd['CSS (µm)']+' '+qxrd.Product


qxrd.loc[qxrd.Composite.isna(),'Composite'] = qxrd.loc[qxrd.CompositeID.isna(),['Mineral']].values
qxrd.rename({'Calcite group - Siderite':'Calcite group Siderite'},inplace=True)
# clean the sizing
data['CSS (µm)'] = data.Size.map(lambda x: clean_size(x))

# create the composite id
data['CompositeID'] = data.BHID.str.strip()+' '+data.Composite.str.strip()+' '+data['CSS (µm)']+' '+data.Type


both = pd.merge(qxrd, data,on=['CompositeID'],how='left')

billet_assay = pd.read_csv(r'reference\billet_assay.csv',low_memory=False)
billet_xrd = pd.read_csv(r'reference\billet_xrd.csv',low_memory=False)
billet = pd.merge( billet_xrd,billet_assay)
billet.rename(columns={'Spinel group - Magnetite':'Spinel group','Calcite group':'Calcite group - Siderite','Hematite':'Hematite group','Calcite group - Calcite':'Calcite group'},inplace=True)
billet['Type'] = 'BILLET'
billet['Product'] = 'BILLET'
billet['Mn'] = billet['MnO']*0.7745
billet.drop(columns=['MnO'],inplace=True)

billet['BHID_x'] = billet.Sample.map(lambda x: x.split('-')[0])
billet.columns = [c.replace(' XRF','') for c in billet.columns]
billet.rename(columns=loi_map, inplace=True)
both = pd.concat([both, billet])
plt.plot(billet['Spinel group'],pd.to_numeric(billet['Fe3O4'],errors='coerce'),'.')
plt.plot(billet['K-Feldspar'],pd.to_numeric(billet['K2O'],errors='coerce'),'.')
plt.show()


plt.plot(both['Spinel group'],pd.to_numeric(both['Fe3O4'],errors='coerce'),'.')
plt.plot(billet['K-Feldspar'],pd.to_numeric(billet['K2O'],errors='coerce'),'.')
plt.show()

keep_columns = [c for c in both.columns if not c.find('_y')>0]
both = both[keep_columns]
both.columns = [c.replace('_x','') for c in both.columns]
both.to_csv(r'reference/BV_XRD_Assay_Merge_3.csv', index=False)
both['PROJECT'] = both.BHID.map(lambda x: x[0:3])
from sklearn.linear_model import RANSACRegressor
coef = []
for i in both.BHID.unique():
    idx= both.BHID == i
    plt.subplot(1, 2, 1)
    plt.plot(both[idx].K2O,both[idx]['Stilpnomelane'],'.',label=i)
    tidx = ~np.isnan(both[idx]['Stilpnomelane'].values.reshape(-1,1))
    lr = RANSACRegressor().fit(both[idx].K2O.values[tidx.ravel()].reshape(-1,1),both[idx]['Stilpnomelane'].values[tidx.ravel()].reshape(-1,1))
    coef.append(lr.estimator_.coef_)
    plt.plot(both.K2O, lr.predict(both.K2O.values.reshape(-1,1)))
    plt.subplot(1, 2, 2)
    plt.plot(both[idx].K2O,both[idx]['K-Feldspar'],'.',label=i)

plt.subplot(1, 2, 1)
plt.legend()
plt.xlabel('K2O')
plt.ylabel('Stilp')

plt.subplot(1, 2, 2)
plt.xlabel('K2O')
plt.ylabel('K-spar')

plt.legend()

plt.show()




